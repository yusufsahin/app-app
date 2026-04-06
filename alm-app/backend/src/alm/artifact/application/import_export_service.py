"""Artifact import/export service helpers."""

from __future__ import annotations

import csv
import io
import json
import uuid
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Any, Literal

from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from alm.area.infrastructure.repositories import SqlAlchemyAreaRepository
from alm.artifact.application.commands.create_artifact import CreateArtifact, CreateArtifactHandler
from alm.artifact.application.commands.update_artifact import UpdateArtifact, UpdateArtifactHandler
from alm.artifact.application.queries.list_artifacts import ListArtifacts, ListArtifactsHandler
from alm.artifact.infrastructure.models import ArtifactModel
from alm.artifact.infrastructure.repositories import SqlAlchemyArtifactRepository
from alm.cycle.infrastructure.repositories import SqlAlchemyCycleRepository
from alm.process_template.infrastructure.repositories import SqlAlchemyProcessTemplateRepository
from alm.project.infrastructure.repositories import SqlAlchemyProjectRepository
from alm.project_tag.infrastructure.repositories import SqlAlchemyProjectTagRepository
from alm.shared.domain.exceptions import ValidationError

ExportFormat = Literal["csv", "xlsx"]
ImportScope = Literal["generic", "testcases", "runs"]
ImportMode = Literal["create", "update", "upsert"]
ImportRowStatus = Literal["created", "updated", "validated", "skipped", "failed"]

CORE_EXPORT_COLUMNS = [
    "artifact_key",
    "artifact_type",
    "title",
    "description",
    "state",
    "parent_key",
    "path",
    "assignee_id",
    "team_id",
    "cycle_id",
    "area_node_id",
    "tag_names",
    "id",
    "parent_id",
    "created_at",
    "updated_at",
]
RUN_EXPORT_COLUMNS = [
    "artifact_key",
    "artifact_type",
    "title",
    "state",
    "path",
    "id",
    "updated_at",
    "cf.environment",
    "cf.run_status_counts_json",
    "cf.expanded_results_json",
]
TESTCASE_STEP_COLUMNS = [
    "test_case_key",
    "step_number",
    "kind",
    "step_id",
    "action",
    "description",
    "expected_result",
    "called_test_case_key",
    "called_test_case_id",
    "called_test_case_title",
    "param_overrides_json",
]
GENERIC_TEMPLATE_EXAMPLE = {
    "artifact_key": "PROJ-101",
    "artifact_type": "requirement",
    "title": "Imported requirement",
    "description": "Created from bulk import",
    "state": "new",
    "parent_key": "",
    "path": "Requirements/Imported requirement",
    "assignee_id": "",
    "team_id": "",
    "cycle_id": "",
    "area_node_id": "",
    "tag_names": "customer;priority-high",
}
TESTCASE_TEMPLATE_EXAMPLE = {
    **GENERIC_TEMPLATE_EXAMPLE,
    "artifact_key": "PROJ-TC-1",
    "artifact_type": "test-case",
    "title": "Verify login succeeds",
    "path": "Quality/Smoke/Verify login succeeds",
}
TESTCASE_STEP_TEMPLATE_EXAMPLES = [
    {
        "test_case_key": "PROJ-TC-1",
        "step_number": "1",
        "kind": "step",
        "step_id": "step-1",
        "action": "Open login page",
        "description": "",
        "expected_result": "Login page is displayed",
        "called_test_case_key": "",
        "called_test_case_id": "",
        "called_test_case_title": "",
        "param_overrides_json": "",
    },
    {
        "test_case_key": "PROJ-TC-1",
        "step_number": "2",
        "kind": "call",
        "step_id": "call-2",
        "action": "",
        "description": "",
        "expected_result": "",
        "called_test_case_key": "PROJ-TC-SHARED",
        "called_test_case_id": "",
        "called_test_case_title": "Shared sign-in flow",
        "param_overrides_json": '{"browser":"chromium"}',
    },
]


@dataclass(slots=True)
class ArtifactExportResult:
    filename: str
    content_type: str
    content: bytes


@dataclass(slots=True)
class ArtifactImportRowResult:
    row_number: int
    sheet: str
    artifact_key: str | None
    status: ImportRowStatus
    message: str | None = None
    artifact_id: uuid.UUID | None = None


@dataclass(slots=True)
class ArtifactImportResult:
    created_count: int = 0
    updated_count: int = 0
    validated_count: int = 0
    skipped_count: int = 0
    failed_count: int = 0
    rows: list[ArtifactImportRowResult] = field(default_factory=list)


@dataclass(slots=True)
class _ImportArtifactRow:
    row_number: int
    sheet: str
    artifact_key: str
    artifact_type: str
    title: str
    description: str
    parent_key: str | None = None
    path: str | None = None
    assignee_id: uuid.UUID | None = None
    team_id: uuid.UUID | None = None
    cycle_id: uuid.UUID | None = None
    area_node_id: uuid.UUID | None = None
    tag_names: list[str] = field(default_factory=list)
    custom_fields: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class _ImportStepRow:
    row_number: int
    sheet: str
    test_case_key: str
    step_number: int
    kind: str
    step_id: str
    action: str = ""
    description: str = ""
    expected_result: str = ""
    called_test_case_key: str | None = None
    called_test_case_id: str | None = None
    called_test_case_title: str | None = None
    param_overrides_json: str | None = None


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, uuid.UUID):
        return str(value)
    return str(value)


def _normalize_header(header: Any) -> str:
    return _stringify(header).strip()


def _parse_uuid(raw: str | None) -> uuid.UUID | None:
    if raw is None:
        return None
    s = raw.strip()
    if not s:
        return None
    return uuid.UUID(s)


def _split_tag_names(raw: str | None) -> list[str]:
    if raw is None:
        return []
    return [part.strip() for part in raw.replace(",", ";").split(";") if part.strip()]


def _coerce_custom_field(raw: str) -> Any:
    text = raw.strip()
    if not text:
        return ""
    if text[0] in "[{\"" or text in {"true", "false", "null"}:
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return raw
    return raw


def _build_path_map(artifacts: list[ArtifactModel]) -> tuple[dict[uuid.UUID, str], dict[uuid.UUID, str | None]]:
    by_id = {artifact.id: artifact for artifact in artifacts}
    cache: dict[uuid.UUID, str] = {}
    parent_keys: dict[uuid.UUID, str | None] = {}

    def resolve_path(artifact_id: uuid.UUID) -> str:
        if artifact_id in cache:
            return cache[artifact_id]
        artifact = by_id[artifact_id]
        parent_key: str | None = None
        if artifact.parent_id and artifact.parent_id in by_id:
            parent_key = by_id[artifact.parent_id].artifact_key
            path = f"{resolve_path(artifact.parent_id)}/{artifact.title}"
        else:
            path = artifact.title
        cache[artifact_id] = path
        parent_keys[artifact_id] = parent_key
        return path

    for artifact in artifacts:
        resolve_path(artifact.id)
    return cache, parent_keys


def _iter_custom_field_keys(artifacts: list[ArtifactModel], *, exclude_test_steps: bool = False) -> list[str]:
    keys: set[str] = set()
    for artifact in artifacts:
        for key in artifact.custom_fields or {}:
            if exclude_test_steps and key == "test_steps_json":
                continue
            keys.add(key)
    return sorted(keys)


def _artifact_to_export_row(
    artifact: ArtifactModel,
    *,
    path: str,
    parent_key: str | None,
    custom_field_keys: list[str],
    run_mode: bool = False,
) -> dict[str, str]:
    row = {
        "artifact_key": artifact.artifact_key or "",
        "artifact_type": artifact.artifact_type,
        "title": artifact.title,
        "description": artifact.description or "",
        "state": artifact.state or "",
        "parent_key": parent_key or "",
        "path": path,
        "assignee_id": _stringify(artifact.assignee_id),
        "team_id": _stringify(artifact.team_id),
        "cycle_id": _stringify(artifact.cycle_id),
        "area_node_id": _stringify(artifact.area_node_id),
        "tag_names": ";".join(tag.name for tag in getattr(artifact, "tags", []) or []),
        "id": _stringify(artifact.id),
        "parent_id": _stringify(artifact.parent_id),
        "created_at": _stringify(artifact.created_at.isoformat() if artifact.created_at else ""),
        "updated_at": _stringify(artifact.updated_at.isoformat() if artifact.updated_at else ""),
    }
    custom_fields = artifact.custom_fields or {}
    for key in custom_field_keys:
        value = custom_fields.get(key)
        if run_mode and key == "run_status_counts_json":
            counts = custom_fields.get("run_status_counts") or {}
            row["cf.run_status_counts_json"] = json.dumps(counts, ensure_ascii=True)
            continue
        if run_mode and key == "expanded_results_json":
            value = custom_fields.get("expanded_results") or custom_fields.get("expandedStepsSnapshot")
            row["cf.expanded_results_json"] = json.dumps(value or [], ensure_ascii=True)
            continue
        row[f"cf.{key}"] = (
            json.dumps(value, ensure_ascii=True) if isinstance(value, (dict, list)) else _stringify(value)
        )
    return row


def _serialize_rows_to_csv(columns: list[str], rows: list[dict[str, str]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({column: row.get(column, "") for column in columns})
    return ("\ufeff" + buffer.getvalue()).encode("utf-8")


def _write_rows_to_sheet(
    workbook: Workbook,
    sheet_name: str,
    columns: list[str],
    rows: list[dict[str, str]],
) -> None:
    ws = workbook.create_sheet(sheet_name)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column, "") for column in columns])


def _parse_test_steps(raw: Any) -> list[dict[str, Any]]:
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except json.JSONDecodeError:
            return []
    if not isinstance(raw, list):
        return []
    parsed: list[dict[str, Any]] = []
    for item in raw:
        if isinstance(item, dict):
            parsed.append(item)
    return parsed


def _testcase_step_rows(artifacts: list[ArtifactModel]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for artifact in artifacts:
        steps = _parse_test_steps((artifact.custom_fields or {}).get("test_steps_json"))
        for index, step in enumerate(steps, start=1):
            kind = _stringify(step.get("kind") or "step")
            rows.append(
                {
                    "test_case_key": artifact.artifact_key or "",
                    "step_number": _stringify(step.get("stepNumber") or index),
                    "kind": kind,
                    "step_id": _stringify(step.get("id") or f"{kind}-{index}"),
                    "action": _stringify(step.get("name")),
                    "description": _stringify(step.get("description")),
                    "expected_result": _stringify(step.get("expectedResult")),
                    "called_test_case_key": _stringify(step.get("calledTestCaseKey")),
                    "called_test_case_id": _stringify(step.get("calledTestCaseId")),
                    "called_test_case_title": _stringify(step.get("calledTitle")),
                    "param_overrides_json": (
                        json.dumps(step.get("paramOverrides") or {}, ensure_ascii=True)
                        if step.get("paramOverrides")
                        else ""
                    ),
                }
            )
    return rows


def _load_csv_rows(payload: bytes) -> list[dict[str, str]]:
    text = payload.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [
        {_normalize_header(k): _stringify(v) for k, v in row.items() if k is not None}
        for row in reader
    ]


def _load_xlsx_rows(payload: bytes) -> dict[str, list[dict[str, str]]]:
    wb = load_workbook(io.BytesIO(payload))
    out: dict[str, list[dict[str, str]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            out[sheet_name] = []
            continue
        headers = [_normalize_header(v) for v in rows[0]]
        parsed: list[dict[str, str]] = []
        for values in rows[1:]:
            row = {
                headers[idx]: _stringify(value)
                for idx, value in enumerate(values)
                if idx < len(headers) and headers[idx]
            }
            if any(cell for cell in row.values()):
                parsed.append(row)
        out[sheet_name] = parsed
    return out


def _load_zip_rows(payload: bytes) -> dict[str, list[dict[str, str]]]:
    out: dict[str, list[dict[str, str]]] = {}
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        for name in archive.namelist():
            if not name.lower().endswith(".csv"):
                continue
            out[name.rsplit("/", 1)[-1].replace(".csv", "")] = _load_csv_rows(archive.read(name))
    return out


def _rows_to_import_artifacts(rows: list[dict[str, str]], *, sheet: str) -> list[_ImportArtifactRow]:
    out: list[_ImportArtifactRow] = []
    for index, row in enumerate(rows, start=2):
        artifact_key = (row.get("artifact_key") or "").strip()
        artifact_type = (row.get("artifact_type") or "").strip()
        title = (row.get("title") or "").strip()
        if not any(row.values()):
            continue
        if not artifact_key:
            raise ValidationError(f"{sheet} row {index}: artifact_key is required")
        if not artifact_type:
            raise ValidationError(f"{sheet} row {index}: artifact_type is required")
        if not title:
            raise ValidationError(f"{sheet} row {index}: title is required")
        custom_fields = {
            key.removeprefix("cf."): _coerce_custom_field(value)
            for key, value in row.items()
            if key.startswith("cf.") and value.strip()
        }
        out.append(
            _ImportArtifactRow(
                row_number=index,
                sheet=sheet,
                artifact_key=artifact_key,
                artifact_type=artifact_type,
                title=title,
                description=(row.get("description") or "").strip(),
                parent_key=(row.get("parent_key") or "").strip() or None,
                path=(row.get("path") or "").strip() or None,
                assignee_id=_parse_uuid(row.get("assignee_id")),
                team_id=_parse_uuid(row.get("team_id")),
                cycle_id=_parse_uuid(row.get("cycle_id")),
                area_node_id=_parse_uuid(row.get("area_node_id")),
                tag_names=_split_tag_names(row.get("tag_names")),
                custom_fields=custom_fields,
            )
        )
    return out


def _rows_to_import_steps(rows: list[dict[str, str]], *, sheet: str) -> list[_ImportStepRow]:
    out: list[_ImportStepRow] = []
    for index, row in enumerate(rows, start=2):
        test_case_key = (row.get("test_case_key") or "").strip()
        if not test_case_key:
            raise ValidationError(f"{sheet} row {index}: test_case_key is required")
        kind = (row.get("kind") or "step").strip() or "step"
        step_number_raw = (row.get("step_number") or "").strip()
        step_number = int(step_number_raw) if step_number_raw else len(out) + 1
        out.append(
            _ImportStepRow(
                row_number=index,
                sheet=sheet,
                test_case_key=test_case_key,
                step_number=step_number,
                kind=kind,
                step_id=(row.get("step_id") or f"{kind}-{step_number}").strip(),
                action=(row.get("action") or "").strip(),
                description=(row.get("description") or "").strip(),
                expected_result=(row.get("expected_result") or "").strip(),
                called_test_case_key=(row.get("called_test_case_key") or "").strip() or None,
                called_test_case_id=(row.get("called_test_case_id") or "").strip() or None,
                called_test_case_title=(row.get("called_test_case_title") or "").strip() or None,
                param_overrides_json=(row.get("param_overrides_json") or "").strip() or None,
            )
        )
    return out


def _build_test_steps_json(
    step_rows: list[_ImportStepRow],
    *,
    known_keys: set[str],
    self_key: str,
) -> list[dict[str, Any]]:
    ordered = sorted(step_rows, key=lambda row: row.step_number)
    out: list[dict[str, Any]] = []
    for step in ordered:
        if step.kind == "call":
            called_key = (step.called_test_case_key or "").strip()
            if not called_key:
                raise ValidationError(
                    f"{step.sheet} row {step.row_number}: called_test_case_key is required for call rows"
                )
            if called_key == self_key:
                    raise ValidationError(
                        f"{step.sheet} row {step.row_number}: test case cannot call itself"
                    )
            if called_key not in known_keys:
                raise ValidationError(
                    f"{step.sheet} row {step.row_number}: called_test_case_key '{called_key}' not found"
                )
            row: dict[str, Any] = {
                "kind": "call",
                "id": step.step_id,
                "stepNumber": step.step_number,
                "calledTestCaseKey": called_key,
            }
            if step.called_test_case_id:
                row["calledTestCaseId"] = step.called_test_case_id
            if step.called_test_case_title:
                row["calledTitle"] = step.called_test_case_title
            if step.param_overrides_json:
                try:
                    parsed = json.loads(step.param_overrides_json)
                except json.JSONDecodeError as exc:
                    raise ValidationError(
                        f"{step.sheet} row {step.row_number}: invalid param_overrides_json"
                    ) from exc
                if not isinstance(parsed, dict):
                    raise ValidationError(
                        f"{step.sheet} row {step.row_number}: param_overrides_json must be an object"
                    )
                row["paramOverrides"] = {str(k): _stringify(v) for k, v in parsed.items()}
            out.append(row)
            continue
        if not step.action:
            raise ValidationError(f"{step.sheet} row {step.row_number}: action is required for step rows")
        out.append(
            {
                "kind": "step",
                "id": step.step_id,
                "stepNumber": step.step_number,
                "name": step.action,
                "description": step.description,
                "expectedResult": step.expected_result,
                "status": "not-executed",
            }
        )
    return out


def _detect_cycle(graph: dict[str, set[str]]) -> str | None:
    visiting: set[str] = set()
    visited: set[str] = set()

    def walk(node: str) -> str | None:
        if node in visited:
            return None
        if node in visiting:
            return node
        visiting.add(node)
        for child in graph.get(node, set()):
            hit = walk(child)
            if hit:
                return hit
        visiting.remove(node)
        visited.add(node)
        return None

    for key in graph:
        hit = walk(key)
        if hit:
            return hit
    return None


async def export_artifacts(
    session: AsyncSession,
    *,
    project_id: uuid.UUID,
    format: ExportFormat,
    scope: ImportScope,
    state: str | None = None,
    type_filter: str | None = None,
    q: str | None = None,
    cycle_id: uuid.UUID | None = None,
    release_id: uuid.UUID | None = None,
    area_node_id: uuid.UUID | None = None,
    sort_by: str | None = None,
    sort_order: str | None = None,
    include_deleted: bool = False,
    include_system_roots: bool = False,
    tree: str | None = None,
    parent_id: uuid.UUID | None = None,
    tag_id: uuid.UUID | None = None,
    team_id: uuid.UUID | None = None,
) -> ArtifactExportResult:
    project_repo = SqlAlchemyProjectRepository(session)
    project = await project_repo.find_by_id(project_id)
    if project is None:
        raise ValidationError("Project not found")
    handler = ListArtifactsHandler(
        artifact_repo=SqlAlchemyArtifactRepository(session),
        project_repo=project_repo,
        cycle_repo=SqlAlchemyCycleRepository(session),
        process_template_repo=SqlAlchemyProcessTemplateRepository(session),
        tag_repo=SqlAlchemyProjectTagRepository(session),
    )
    query_result = await handler.handle(
        ListArtifacts(
            tenant_id=project.tenant_id,
            project_id=project_id,
            state_filter=state,
            type_filter=type_filter,
            search_query=q,
            cycle_id=cycle_id,
            release_id=release_id,
            area_node_id=area_node_id,
            sort_by=sort_by,
            sort_order=sort_order,
            limit=None,
            offset=None,
            include_deleted=include_deleted,
            include_system_roots=include_system_roots,
            tree=tree,
            parent_id=parent_id,
            tag_id=tag_id,
            team_id=team_id,
        )
    )
    models = query_result.items
    path_map, parent_key_map = _build_path_map(models)

    if scope == "runs":
        custom_field_keys = ["environment", "run_status_counts_json", "expanded_results_json"]
        rows = [
            _artifact_to_export_row(
                artifact,
                path=path_map[artifact.id],
                parent_key=parent_key_map.get(artifact.id),
                custom_field_keys=custom_field_keys,
                run_mode=True,
            )
            for artifact in models
        ]
        columns = RUN_EXPORT_COLUMNS
        filename_base = "runs"
    elif scope == "testcases":
        custom_field_keys = _iter_custom_field_keys(models, exclude_test_steps=True)
        artifact_rows = [
            _artifact_to_export_row(
                artifact,
                path=path_map[artifact.id],
                parent_key=parent_key_map.get(artifact.id),
                custom_field_keys=custom_field_keys,
            )
            for artifact in models
        ]
        step_rows = _testcase_step_rows(models)
        if format == "csv":
            payload = io.BytesIO()
            with zipfile.ZipFile(payload, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                artifact_columns = CORE_EXPORT_COLUMNS + [f"cf.{k}" for k in custom_field_keys]
                archive.writestr("artifacts.csv", _serialize_rows_to_csv(artifact_columns, artifact_rows))
                archive.writestr("test_case_steps.csv", _serialize_rows_to_csv(TESTCASE_STEP_COLUMNS, step_rows))
            return ArtifactExportResult(
                filename="testcases-export.zip",
                content_type="application/zip",
                content=payload.getvalue(),
            )
        workbook = Workbook()
        workbook.remove(workbook.active)
        artifact_columns = CORE_EXPORT_COLUMNS + [f"cf.{k}" for k in custom_field_keys]
        _write_rows_to_sheet(workbook, "artifacts", artifact_columns, artifact_rows)
        _write_rows_to_sheet(workbook, "test_case_steps", TESTCASE_STEP_COLUMNS, step_rows)
        buffer = io.BytesIO()
        workbook.save(buffer)
        return ArtifactExportResult(
            filename="testcases-export.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content=buffer.getvalue(),
        )
    else:
        custom_field_keys = _iter_custom_field_keys(models)
        rows = [
            _artifact_to_export_row(
                artifact,
                path=path_map[artifact.id],
                parent_key=parent_key_map.get(artifact.id),
                custom_field_keys=custom_field_keys,
            )
            for artifact in models
        ]
        columns = CORE_EXPORT_COLUMNS + [f"cf.{k}" for k in custom_field_keys]
        filename_base = "artifacts"

    if format == "csv":
        return ArtifactExportResult(
            filename=f"{filename_base}-export.csv",
            content_type="text/csv; charset=utf-8",
            content=_serialize_rows_to_csv(columns, rows),
        )
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_rows_to_sheet(workbook, "artifacts", columns, rows)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return ArtifactExportResult(
        filename=f"{filename_base}-export.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=buffer.getvalue(),
    )


async def export_import_template(format: ExportFormat, scope: ImportScope) -> ArtifactExportResult:
    if scope == "testcases":
        artifact_rows = [{**TESTCASE_TEMPLATE_EXAMPLE}]
        step_rows = list(TESTCASE_STEP_TEMPLATE_EXAMPLES)
        if format == "csv":
            payload = io.BytesIO()
            with zipfile.ZipFile(payload, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                archive.writestr("artifacts.csv", _serialize_rows_to_csv(CORE_EXPORT_COLUMNS, artifact_rows))
                archive.writestr("test_case_steps.csv", _serialize_rows_to_csv(TESTCASE_STEP_COLUMNS, step_rows))
            return ArtifactExportResult(
                filename="artifact-import-template-testcases.zip",
                content_type="application/zip",
                content=payload.getvalue(),
            )
        workbook = Workbook()
        workbook.remove(workbook.active)
        _write_rows_to_sheet(workbook, "artifacts", CORE_EXPORT_COLUMNS, artifact_rows)
        _write_rows_to_sheet(workbook, "test_case_steps", TESTCASE_STEP_COLUMNS, step_rows)
        buffer = io.BytesIO()
        workbook.save(buffer)
        return ArtifactExportResult(
            filename="artifact-import-template-testcases.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content=buffer.getvalue(),
        )
    rows = [GENERIC_TEMPLATE_EXAMPLE]
    if format == "csv":
        return ArtifactExportResult(
            filename="artifact-import-template.csv",
            content_type="text/csv; charset=utf-8",
            content=_serialize_rows_to_csv(CORE_EXPORT_COLUMNS, rows),
        )
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_rows_to_sheet(workbook, "artifacts", CORE_EXPORT_COLUMNS, rows)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return ArtifactExportResult(
        filename="artifact-import-template.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=buffer.getvalue(),
    )


def _parse_import_payload(
    payload: bytes,
    filename: str,
    scope: ImportScope,
) -> tuple[list[_ImportArtifactRow], list[_ImportStepRow]]:
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        sheets = _load_xlsx_rows(payload)
    elif lower.endswith(".zip"):
        sheets = _load_zip_rows(payload)
    else:
        sheets = {"artifacts": _load_csv_rows(payload)}
    artifact_rows = _rows_to_import_artifacts(
        sheets.get("artifacts", []),
        sheet="artifacts",
    )
    step_rows: list[_ImportStepRow] = []
    if scope == "testcases":
        step_rows = _rows_to_import_steps(
            sheets.get("test_case_steps", []),
            sheet="test_case_steps",
        )
    return artifact_rows, step_rows


async def import_artifacts(
    session: AsyncSession,
    *,
    tenant_id: uuid.UUID,
    project_id: uuid.UUID,
    actor_id: uuid.UUID | None,
    filename: str,
    payload: bytes,
    scope: ImportScope,
    mode: ImportMode,
    validate_only: bool,
) -> ArtifactImportResult:
    artifact_repo = SqlAlchemyArtifactRepository(session)
    project_repo = SqlAlchemyProjectRepository(session)
    area_repo = SqlAlchemyAreaRepository(session)
    process_template_repo = SqlAlchemyProcessTemplateRepository(session)
    tag_repo = SqlAlchemyProjectTagRepository(session)
    create_handler = CreateArtifactHandler(
        artifact_repo=artifact_repo,
        project_repo=project_repo,
        process_template_repo=process_template_repo,
        area_repo=area_repo,
        tag_repo=tag_repo,
    )
    update_handler = UpdateArtifactHandler(
        artifact_repo=artifact_repo,
        project_repo=project_repo,
        area_repo=area_repo,
        process_template_repo=process_template_repo,
        tag_repo=tag_repo,
    )
    artifacts, step_rows = _parse_import_payload(payload, filename, scope)
    existing = await artifact_repo.list_by_project(
        project_id=project_id,
        limit=None,
        offset=None,
        include_deleted=False,
    )
    existing_by_key = {artifact.artifact_key: artifact for artifact in existing if artifact.artifact_key}
    paths, _ = _build_path_map(existing)
    existing_path_to_id = {path: artifact_id for artifact_id, path in paths.items()}
    step_rows_by_key: dict[str, list[_ImportStepRow]] = defaultdict(list)
    for step_row in step_rows:
        step_rows_by_key[step_row.test_case_key].append(step_row)
    graph: dict[str, set[str]] = defaultdict(set)
    for test_case_key, grouped in step_rows_by_key.items():
        for step in grouped:
            if step.kind == "call" and step.called_test_case_key:
                graph[test_case_key].add(step.called_test_case_key)
    if cycle := _detect_cycle(graph):
        raise ValidationError(f"Detected circular test-case call graph involving '{cycle}'")

    pending = list(artifacts)
    known_keys = {artifact.artifact_key for artifact in artifacts} | set(existing_by_key.keys())
    result = ArtifactImportResult()

    while pending:
        progressed = False
        next_pending: list[_ImportArtifactRow] = []
        for row in pending:
            try:
                current = existing_by_key.get(row.artifact_key)
                if mode == "create" and current is not None:
                    raise ValidationError(
                        f"{row.sheet} row {row.row_number}: artifact_key '{row.artifact_key}' already exists"
                    )
                if mode == "update" and current is None:
                    raise ValidationError(
                        f"{row.sheet} row {row.row_number}: artifact_key '{row.artifact_key}' was not found"
                    )

                parent_id: uuid.UUID | None = None
                if row.parent_key:
                    parent = existing_by_key.get(row.parent_key)
                    if parent is None:
                        sibling = next((item for item in artifacts if item.artifact_key == row.parent_key), None)
                        if sibling is not None:
                            next_pending.append(row)
                            continue
                        raise ValidationError(
                            f"{row.sheet} row {row.row_number}: parent_key '{row.parent_key}' was not found"
                        )
                    parent_id = parent.id
                elif row.path and "/" in row.path:
                    parent_path = row.path.rsplit("/", 1)[0]
                    parent_id = existing_path_to_id.get(parent_path)
                    if parent_id is None and parent_path:
                        next_pending.append(row)
                        continue

                custom_fields = dict(row.custom_fields)
                if scope == "testcases" and row.artifact_type == "test-case":
                    custom_fields["test_steps_json"] = _build_test_steps_json(
                        step_rows_by_key.get(row.artifact_key, []),
                        known_keys=known_keys,
                        self_key=row.artifact_key,
                    )

                if current is None:
                    dto = await create_handler.handle(
                        CreateArtifact(
                            tenant_id=tenant_id,
                            project_id=project_id,
                            artifact_type=row.artifact_type,
                            title=row.title,
                            description=row.description,
                            parent_id=parent_id,
                            assignee_id=row.assignee_id,
                            custom_fields=custom_fields,
                            artifact_key=row.artifact_key,
                            cycle_id=row.cycle_id,
                            area_node_id=row.area_node_id,
                            team_id=row.team_id,
                            created_by=actor_id,
                        )
                    )
                    entity = await artifact_repo.find_by_id(dto.id)
                    if entity is not None:
                        existing_by_key[row.artifact_key] = entity
                        paths, _ = _build_path_map(list(existing_by_key.values()))
                        existing_path_to_id = {path: artifact_id for artifact_id, path in paths.items()}
                    status: ImportRowStatus = "validated" if validate_only else "created"
                    if validate_only:
                        result.validated_count += 1
                    else:
                        result.created_count += 1
                    result.rows.append(
                        ArtifactImportRowResult(
                            row_number=row.row_number,
                            sheet=row.sheet,
                            artifact_key=row.artifact_key,
                            status=status,
                            artifact_id=dto.id,
                        )
                    )
                else:
                    update_payload: dict[str, Any] = {
                        "title": row.title,
                        "description": row.description,
                        "assignee_id": row.assignee_id,
                        "custom_fields": custom_fields,
                        "cycle_id": row.cycle_id,
                        "area_node_id": row.area_node_id,
                        "team_id": row.team_id,
                    }
                    # Omit parent_id when the row does not specify a parent; passing None would clear the parent,
                    # which is invalid for non-root artifacts.
                    if row.parent_key or (row.path and "/" in row.path):
                        update_payload["parent_id"] = parent_id
                    dto = await update_handler.handle(
                        UpdateArtifact(
                            tenant_id=tenant_id,
                            project_id=project_id,
                            artifact_id=current.id,
                            updates=update_payload,
                            updated_by=actor_id,
                        )
                    )
                    entity = await artifact_repo.find_by_id(dto.id)
                    if entity is not None:
                        existing_by_key[row.artifact_key] = entity
                        paths, _ = _build_path_map(list(existing_by_key.values()))
                        existing_path_to_id = {path: artifact_id for artifact_id, path in paths.items()}
                    status = "validated" if validate_only else "updated"
                    if validate_only:
                        result.validated_count += 1
                    else:
                        result.updated_count += 1
                    result.rows.append(
                        ArtifactImportRowResult(
                            row_number=row.row_number,
                            sheet=row.sheet,
                            artifact_key=row.artifact_key,
                            status=status,
                            artifact_id=dto.id,
                        )
                    )
                progressed = True
            except Exception as exc:  # noqa: BLE001
                result.failed_count += 1
                result.rows.append(
                    ArtifactImportRowResult(
                        row_number=row.row_number,
                        sheet=row.sheet,
                        artifact_key=row.artifact_key,
                        status="failed",
                        message=str(exc),
                    )
                )
        if not progressed and next_pending:
            for row in next_pending:
                result.failed_count += 1
                result.rows.append(
                    ArtifactImportRowResult(
                        row_number=row.row_number,
                        sheet=row.sheet,
                        artifact_key=row.artifact_key,
                        status="failed",
                        message="Parent dependency could not be resolved from this import batch",
                    )
                )
            break
        pending = next_pending

    if validate_only:
        await session.rollback()
    return result
